"""
Модуль для экспорта данных в Excel
"""
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_users_to_excel(users: list) -> bytes:
    """
    Экспортирует список пользователей в Excel-файл (для админа)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        "ID пользователя",
        "Имя",
        "Username",
        "Дата регистрации",
        "Статус подписки",
        "Дата окончания"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, user in enumerate(users, 2):
        user_id = user.get("user_id", "")
        first_name = user.get("first_name", "")
        username = f"@{user['username']}" if user.get("username") else ""
        
        created_at = user.get("created_at", "")
        if created_at and len(created_at) >= 10:
            parts = created_at[:10].split('-')
            if len(parts) == 3:
                created_at = f"{parts[2]}.{parts[1]}.{parts[0]}"
        
        status = ""
        end_date = ""
        
        if user.get("is_forever"):
            status = "Бессрочная"
            end_date = "∞"
        elif user.get("paid_until"):
            status = "Оплачена"
            end_date = user["paid_until"]
            if end_date and len(end_date) >= 10:
                parts = end_date[:10].split('-')
                if len(parts) == 3:
                    end_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
        elif user.get("trial_end"):
            status = "Триал"
            end_date = user["trial_end"]
            if end_date and len(end_date) >= 10:
                parts = end_date[:10].split('-')
                if len(parts) == 3:
                    end_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
        else:
            status = "Не активна"
            end_date = "-"
        
        row_data = [user_id, first_name, username, created_at, status, end_date]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    ws_stats = wb.create_sheet("Статистика")
    
    total_users = len(users)
    active_users = sum(1 for u in users if (
        u.get("is_forever") or 
        (u.get("paid_until") and u["paid_until"] >= datetime.now().strftime("%Y-%m-%d")) or
        (u.get("trial_end") and u["trial_end"] >= datetime.now().strftime("%Y-%m-%d"))
    ))
    trial_users = sum(1 for u in users if u.get("trial_end") and not u.get("paid_until") and not u.get("is_forever"))
    paid_users = sum(1 for u in users if u.get("paid_until") or u.get("is_forever"))
    
    stats_data = [
        ["Показатель", "Значение"],
        ["Всего пользователей", total_users],
        ["Активных пользователей", active_users],
        ["На триале", trial_users],
        ["Оплативших", paid_users],
        ["Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")]
    ]
    
    for row_idx, row in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
    
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_user_meals_to_excel(meals: list, days: int = 30) -> bytes:
    """
    Экспорт приёмов пищи пользователя в Excel
    """
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============ Лист 1: Журнал питания ============
    ws = wb.active
    ws.title = "Журнал питания"
    
    headers = ["Дата", "Время", "Продукт", "Вес (г)", "Ккал", "Белки", "Жиры", "Углеводы"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, meal in enumerate(meals, 2):
        meal_time = meal.get("meal_time", "")
        date_str = ""
        time_str = ""
        if meal_time:
            parts = meal_time.split()
            if len(parts) >= 1:
                date_parts = parts[0].split('-')
                if len(date_parts) == 3:
                    date_str = f"{date_parts[2]}.{date_parts[1]}.{date_parts[0]}"
            if len(parts) >= 2:
                time_str = parts[1][:5]
        
        row_data = [
            date_str,
            time_str,
            meal.get("product_name", ""),
            meal.get("weight_grams", 0),
            round(meal.get("calories", 0), 1),
            round(meal.get("protein", 0), 1),
            round(meal.get("fat", 0), 1),
            round(meal.get("carbohydrates", 0), 1)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # ============ Лист 2: Сводка по дням ============
    ws_days = wb.create_sheet("Сводка по дням")
    
    day_headers = ["Дата", "Приёмов", "Ккал", "Белки", "Жиры", "Углеводы"]
    
    for col, header in enumerate(day_headers, 1):
        cell = ws_days.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    days_summary = {}
    for meal in meals:
        meal_time = meal.get("meal_time", "")
        if meal_time:
            day = meal_time[:10]
            if day not in days_summary:
                days_summary[day] = {"count": 0, "calories": 0, "protein": 0, "fat": 0, "carbs": 0}
            days_summary[day]["count"] += 1
            days_summary[day]["calories"] += meal.get("calories", 0)
            days_summary[day]["protein"] += meal.get("protein", 0)
            days_summary[day]["fat"] += meal.get("fat", 0)
            days_summary[day]["carbs"] += meal.get("carbohydrates", 0)
    
    row_idx = 2
    for day in sorted(days_summary.keys(), reverse=True):
        data = days_summary[day]
        parts = day.split('-')
        day_formatted = f"{parts[2]}.{parts[1]}.{parts[0]}" if len(parts) == 3 else day
        
        row_data = [
            day_formatted,
            data["count"],
            round(data["calories"], 1),
            round(data["protein"], 1),
            round(data["fat"], 1),
            round(data["carbs"], 1)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_days.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
        row_idx += 1
    
    for col in range(1, len(day_headers) + 1):
        ws_days.column_dimensions[get_column_letter(col)].width = 15
    
    # ============ Лист 3: Общая статистика ============
    ws_stats = wb.create_sheet("Общая статистика")
    
    total_days = len(days_summary)
    total_meals = len(meals)
    
    avg_calories = sum(d["calories"] for d in days_summary.values()) / max(total_days, 1)
    avg_protein = sum(d["protein"] for d in days_summary.values()) / max(total_days, 1)
    avg_fat = sum(d["fat"] for d in days_summary.values()) / max(total_days, 1)
    avg_carbs = sum(d["carbs"] for d in days_summary.values()) / max(total_days, 1)
    
    stats_data = [
        ["Показатель", "Значение"],
        ["Период", f"{days} дней"],
        ["Всего дней с записями", total_days],
        ["Всего приёмов пищи", total_meals],
        ["", ""],
        ["Среднее в день:", ""],
        ["Калории", f"{avg_calories:.0f} ккал"],
        ["Белки", f"{avg_protein:.1f} г"],
        ["Жиры", f"{avg_fat:.1f} г"],
        ["Углеводы", f"{avg_carbs:.1f} г"],
        ["", ""],
        ["Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")]
    ]
    
    for row_idx, row in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
    
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()